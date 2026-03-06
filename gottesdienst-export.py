import requests
import locale
import os
import sys
import calendar
from datetime import date, timedelta, datetime
from zoneinfo import ZoneInfo
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
from config import BASE_URL, use_email, smtp_server, smtp_port, smtp_username, email_recipients
#use_email = False

if use_email:
    # Import smtplib for the actual sending function
    import smtplib

    # Here are the email package modules we'll need
    from email.message import EmailMessage

# Prüfe, ob mindestens Python 3.6 installiert ist
# Diese Version ist notwendig, um die Datetime- und Dateutil-Bibliotheken und f-Strings zu verwenden.
if sys.version_info < (3, 6):
    raise Exception("Dieses Skript benötigt Python 3.6 oder höher.")

# Lese die SMTP Konfiguration aus der Umgebungsvariablen
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")

if use_email and not SMTP_PASSWORD:
    raise ValueError("Das SMTP Passwort ist nicht gesetzt. Bitte die Umgebungsvariable 'SMTP_PASSWORD' definieren.")

# Konfigurationsparameter für verschiedene ChurchTools Instanzen
parameters = [
    {
        "name" : "Volksdorf",
        "id" : "VOL",
        "calendar_ids" : [2],
        "calender_title" : "Gottesdienste",
        "url_name" : "volksdorf",
        "resource_ids" : [3, 5, 9, 10, 11, 12, 13, 14, 22],
        "musik_service_id" : 2,
        "pastores_service_id" : 3
    },
    {
        "name" : "Duvenstedt",
        "id" : "DUV",
        "calendar_ids" : [2],
        "calender_title" : "Gottesdienst (durch PastorInnen zu erstellen)",
        "url_name" : "cantatekirche",
        "resource_ids" : [8],
        "musik_service_id" : 2,
        "pastores_service_id" : 3
    },
    {
        "name" : "Oberalster Bergstedt",
        "id" : "OAB",
        "calendar_ids" : [35, 38, 41],
        "calender_title" : "Gottesdienst",
        "url_name" : "oberalster-bergstedt",
        "resource_ids" : [8, 11, 12, 15, 19, 20, 28, 29, 30, 32, 33, 34, 39, 41, 42],
        "musik_service_id" : 10,
        "pastores_service_id" : 1
    }
]
FROM = date.today().strftime("%Y-%m-%d")
one_year_from_today = date.today() + timedelta(days=365)
TO = one_year_from_today.strftime("%Y-%m-%d")

def get_services(from_date=None, to_date=None):
    _from = from_date if from_date else FROM
    _to   = to_date   if to_date   else TO
    rows = [
        ['Gemeinde', 'Standort', 'Datum (Sortierung)', 'Datum', 'Beginn', 'Ende', 'Name',
         'Dienstplan_Notiz', 'Kalender_Untertitel', 'Kalender_Beschreibung', 'Raum', 'Pastor*in', 'Organist*in'],
    ]
    for param in parameters:
        ID = param.get("id")
        NAME = param.get("name")
        BASE_URL = f"https://{param.get("url_name")}.church.tools/api"
        CALENDAR_IDS = param.get("calendar_ids")
        CALENDAR_TITLE = param.get("calender_title")
        RESOURCE_IDS = param.get("resource_ids")
        MUSIK_SERVICE_ID = param.get("musik_service_id")
        PASTORES_SERVICE_ID = param.get("pastores_service_id")

        # Lese das Zugangstoken aus der Umgebungsvariablen
        TOKEN = os.getenv(f"CHURCHTOOLS_TOKEN_{ID}")

        if not TOKEN:
            print(f"Das Access Token ist nicht gesetzt. Bitte die Umgebungsvariable 'CHURCHTOOLS_TOKEN_{ID}' definieren.")
            continue

        # Setze die Locale auf Deutsch, damit der ausgeschriebene Monatsname in den Datumsangaben auf Deutsch ist
        locale.setlocale(locale.LC_TIME, "de_DE.utf8")

        # Lade alle Gottesdienste im Events Modul (Dienstplan)
        # -> events
        headers = {
            "Authorization": f"Login {TOKEN}"
        }
        params = {
            # Kalender ID 2 ist der Gottesdienste Kalender, wir nehmen nur Events aus diesem Kalender
            "calendar_ids[]": CALENDAR_IDS,
            "from": _from,
            # Das "to" Feld ist nötig, weil sonst nur Events des aktuellen Monats geladen werden
            "to": _to,
            # Lade auch Details zu Diensten wie Musik und Pastor*in
            "include": "eventServices",
            # Keine abgesagten Events
            "canceled": "false",
        }

        events_response = requests.get(f"{BASE_URL}/events", headers=headers, params=params)    
        events_response.raise_for_status()
        events = events_response.json().get("data", [])
        pagination = events_response.json().get("meta").get("pagination")
        if pagination:
            pages = pagination.get("lastPage")
            for page in range(2, pages + 1):
                events_response = requests.get(f"{BASE_URL}/events?page={page}", headers=headers, params=params)
                events_response.raise_for_status()
                events.extend(events_response.json().get("data", []))
        print(f"{len(events)} Events für {NAME} gefunden.")
        # Lade alle Einträge aus dem Gottesdienste Kalender
        # -> appointments
        params = {
            "from": _from,
            # Das "to" Feld ist nötig, weil sonst nur Termine des aktuellen Monats geladen werden
            "to": _to,
        }
        # Gottesdienste Kalender hat die ID 2 für Volksdorf
        appointments_response = requests.get(f"{BASE_URL}/calendars/{CALENDAR_IDS[0]}/appointments", headers=headers, params=params)    
        appointments_response.raise_for_status()
        appointments = appointments_response.json().get("data", [])
        pagination = appointments_response.json().get("meta").get("pagination")
        if pagination:
            pages = pagination.get("lastPage")
            for page in range(2, pages + 1):
                appointments_response = requests.get(f"{BASE_URL}/calendars/{CALENDAR_IDS[0]}/appointments?page={page}", headers=headers, params=params)
                appointments_response.raise_for_status()
                appointments.extend(appointments_response.json().get("data", []))
        print(f"{len(appointments)} Kalendereinträge für {NAME} gefunden.")

        # Lade alle Raumbuchungen für die Gottesdienste
        params = {
            # Gottesdienst Räume
            "resource_ids[]": RESOURCE_IDS,
            "from": _from,
            # Das "to" Feld ist nötig, weil sonst nur Buchungen des aktuellen Monats geladen werden
            "to": _to,
        }
        bookings_response = requests.get(f"{BASE_URL}/bookings", headers=headers, params=params)

        bookings_response.raise_for_status()
        bookings = bookings_response.json().get("data", [])
        pagination = bookings_response.json().get("meta").get("pagination")
        if pagination:
            pages = pagination.get("lastPage")
            for page in range(2, pages + 1):
                bookings_response = requests.get(f"{BASE_URL}/bookings?page={page}", headers=headers, params=params)
                bookings_response.raise_for_status()
                bookings.extend(bookings_response.json().get("data", []))
        print(f"{len(bookings)} Raumbuchungen für {NAME} gefunden.")

        # Erstelle die CSV Datei mit den Gottesdiensten
        # CSV Spalten:
        # Gemeinde               - 'Volksdorf' (statisch)
        # Datum                  - 'Montag, den 01.01.2024' (aus event.startDate)
        # Beginn                 - '10:00' (aus event.startDate)
        # Ende                   - '12:00' (aus event.endDate)
        # Name                   - Gottesdienst (aus event.name)
        # Dienstplan_Notiz       - Notiz zum Dienstplan (aus event.note)
        # Kalender_Untertitel    - Untertitel des Kalendereintrags (aus appointment.base.subtitle)
        # Kalender_Beschreibung  - Beschreibung des Kalendereintrags (aus appointment.base.description)
        # Raum                   - Raum/Kirche (aus booking.base.resource.name)
        # Pastor*in              - Pastor*in (aus event.eventServices mit serviceId 3)                                               
        # Organist*in            - Organist*in (aus event.eventServices mit serviceId 2)
        # Gottesdienste
        for event in events:
            if event.get("calendar").get("title") == CALENDAR_TITLE:
                if event.get("calendar").get("domainAttributes").get("campusName"):
                    location = event.get("calendar").get("domainAttributes").get("campusName")
                else:
                    location = "(kein Standort)"
                eventServices = event.get("eventServices", [])
                event_start_date = datetime.fromisoformat(event.get("startDate")).astimezone(_BERLIN)
                event_end_date = datetime.fromisoformat(event.get("endDate")).astimezone(_BERLIN)
                pastores = []
                musik = []
                for service in eventServices:
                    if service.get("name"):
                        if service.get("serviceId") == MUSIK_SERVICE_ID:
                            musik.append(service.get("name"))
                        if service.get("serviceId") == PASTORES_SERVICE_ID:
                            pastores.append(service.get("name"))
                appointment_id = event.get("appointmentId")
                appointment = next((a for a in appointments if a.get("base").get("id") == appointment_id), None)
                if appointment:
                    ab = appointment.get('base')
                    subtitle = ab.get('subtitle') or '(kein Untertitel)'
                    description = ab.get('description') or '(keine Beschreibung)'
                else:
                    subtitle = '(Kein Kalendereintrag)'
                    description = ''
                booking = next((b for b in bookings if b.get("base").get("appointmentId") == appointment_id), None)
                if booking:
                    b_text = booking.get('base').get('resource').get('name')
                else:
                    b_text = "(Keine Raumbuchung)"
                note = event.get('note')
                if not note:
                    note = '(keine Dienstplan-Notiz)'
                if not pastores:
                    pastores.append('(keine Pastor*in)')
                if not musik:
                    musik.append('(kein Organist*in)')
                rows.append([
                    NAME,
                    location,
                    event_start_date.replace(tzinfo=None),
                    f"{calendar.day_name[event_start_date.weekday()]}, den {event_start_date.strftime('%d.%m.%Y')}",
                    event_start_date.strftime('%H:%M'),
                    event_end_date.strftime('%H:%M'),
                    event.get('name'),
                    note,
                    subtitle,
                    description,
                    b_text,
                    " | ".join(pastores),
                    " | ".join(musik),
                ])

    # Kirchenjahr – calendar ID 14, Volksdorf only
    kirchenjahr_rows = []
    vol_param = next((p for p in parameters if p.get("id") == "VOL"), None)
    TOKEN_VOL = os.getenv("CHURCHTOOLS_TOKEN_VOL")
    if vol_param and TOKEN_VOL:
        BASE_URL_VOL = f"https://{vol_param.get('url_name')}.church.tools/api"
        headers = {"Authorization": f"Login {TOKEN_VOL}"}
        params = {"from": _from, "to": _to}
        kj_response = requests.get(f"{BASE_URL_VOL}/calendars/14/appointments", headers=headers, params=params)
        kj_response.raise_for_status()
        kj_appointments = kj_response.json().get("data", [])
        pagination = kj_response.json().get("meta", {}).get("pagination")
        if pagination:
            pages = pagination.get("lastPage")
            for page in range(2, pages + 1):
                kj_response = requests.get(f"{BASE_URL_VOL}/calendars/14/appointments?page={page}", headers=headers, params=params)
                kj_response.raise_for_status()
                kj_appointments.extend(kj_response.json().get("data", []))
        print(f"{len(kj_appointments)} Kirchenjahr-Einträge gefunden.")
        for appt in sorted(kj_appointments, key=lambda a: a.get("base", {}).get("startDate", "")):
            base = appt.get("base", {})
            start_date = base.get("startDate")
            caption = base.get("caption", "")
            if start_date:
                kirchenjahr_rows.append([datetime.fromisoformat(start_date).astimezone(_BERLIN).replace(tzinfo=None), caption])

    return rows, kirchenjahr_rows

_BERLIN = ZoneInfo("Europe/Berlin")

ARIAL = Font(name="Arial")
ARIAL_BOLD = Font(name="Arial", bold=True)
ARIAL_BOLD_LARGE = Font(name="Arial", bold=True, size=12)
LIGHT_BLUE_FILL = PatternFill(start_color="DAEAF6", end_color="DAEAF6", fill_type="solid")
_THIN = Side(style='thin')
SERVICE_ROW_BORDER = Border(top=_THIN, bottom=_THIN)
_CHARS_PER_LINE_C = 60   # approx characters per line in col C at 11 cm, Arial 11pt
_LINE_HEIGHT_PT   = 14   # Excel line height for Arial 11pt
_ROW_PADDING_PT   = 12   # 6pt top + 6pt bottom

def _build_gemeindebrief_sheet(wb, sheet_title, sorted_data, highlight_gemeinde, kirchenjahr_by_date):
    # Columns in data rows:
    #  0=Gemeinde, 1=Standort, 2=Datum(sort), 3=Datum(text), 4=Beginn,
    #  5=Ende, 6=Name, 7=Notiz, 8=Untertitel, 9=Beschreibung,
    #  10=Raum, 11=Pastor*in, 12=Organist*in
    ws = wb.create_sheet(title=sheet_title)
    ws.sheet_view.showGridLines = False
    current_date = None
    row_idx = 1
    for service in sorted_data:
        dt = service[2]  # datetime object
        date_key = dt.date()
        if date_key != current_date:
            current_date = date_key
            kj_title = kirchenjahr_by_date.get(date_key)
            bracket_text = kj_title if kj_title else calendar.day_name[dt.weekday()]
            date_heading = f"{dt.strftime('%d.%m.%Y')} ({bracket_text})"
            cell = ws.cell(row=row_idx, column=1, value=date_heading)
            cell.font = ARIAL_BOLD_LARGE
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
            ws.row_dimensions[row_idx].height = 30  # double the default (~15pt)
            row_idx += 1
            ws.row_dimensions[row_idx].height = 6   # 6pt spacer below date heading
            row_idx += 1
        name       = service[6] or ''
        untertitel = service[8] or ''
        pastor     = service[11] or ''
        organist   = service[12] or ''
        desc_parts = [name]
        if untertitel not in ('(kein Untertitel)', '(Kein Kalendereintrag)'):
            desc_parts.append(untertitel)
        if pastor != '(keine Pastor*in)':
            desc_parts.append(f"P: {pastor}")
        if organist != '(kein Organist*in)':
            desc_parts.append(f"Musik: {organist}")
        description = ' - '.join(desc_parts)
        fill = LIGHT_BLUE_FILL if service[0] == highlight_gemeinde else None
        # Build column B: standort bold; for Volksdorf append cleaned room in normal weight
        standort = service[1]
        raum = service[10]
        if service[0] == "Volksdorf" and raum != "(Keine Raumbuchung)":
            raum_clean = raum.replace("Kirche ", "").strip()
            if raum_clean:
                col_b = CellRichText([
                    TextBlock(InlineFont(b=True,  rFont="Arial"), standort),
                    TextBlock(InlineFont(b=False, rFont="Arial"), f" ({raum_clean})"),
                ])
            else:
                col_b = CellRichText([TextBlock(InlineFont(b=True, rFont="Arial"), standort)])
        else:
            col_b = CellRichText([TextBlock(InlineFont(b=True, rFont="Arial"), standort)])
        for col, value in enumerate([service[4], col_b, description], start=1):
            c = ws.cell(row=row_idx, column=col, value=value)
            if col != 2:
                c.font = ARIAL  # col 2 font is embedded in CellRichText runs
            c.alignment = Alignment(vertical='center', wrap_text=(col >= 2))
            c.border = SERVICE_ROW_BORDER
            if fill:
                c.fill = fill
        n_lines = max(1, -(-len(description) // _CHARS_PER_LINE_C))
        ws.row_dimensions[row_idx].height = n_lines * _LINE_HEIGHT_PT + _ROW_PADDING_PT
        row_idx += 1
    # Column A: fixed width just for "00:00"
    ws.column_dimensions['A'].width = 7
    # Column B: 5 cm (≈ 25 Excel character units)
    ws.column_dimensions['B'].width = 25
    # Column C: 11 cm (≈ 55 Excel character units)
    ws.column_dimensions['C'].width = 55

def create_excel(rows, kirchenjahr_rows):
    wb = openpyxl.Workbook()

    # Sheet 1: Gottesdienste (full data)
    ws1 = wb.active
    ws1.title = "Gottesdienste"
    for i, row in enumerate(rows):
        ws1.append(row)
        if i == 0:
            for cell in ws1[1]:
                cell.font = ARIAL_BOLD
        else:
            for cell in ws1[i + 1]:
                cell.font = ARIAL

    # Sheets 2–4: Gemeindebrief per Gemeinde
    sorted_data = sorted(rows[1:], key=lambda r: r[2])  # sort by datetime (skip 1 header row)
    kirchenjahr_by_date = {dt.date(): caption for dt, caption in kirchenjahr_rows}
    _build_gemeindebrief_sheet(wb, "Gemeindebrief - Volksdorf",             sorted_data, "Volksdorf",            kirchenjahr_by_date)
    _build_gemeindebrief_sheet(wb, "Gemeindebrief - Duvenstedt",            sorted_data, "Duvenstedt",           kirchenjahr_by_date)
    _build_gemeindebrief_sheet(wb, "Gemeindebrief - OA-Bergstedt",          sorted_data, "Oberalster Bergstedt", kirchenjahr_by_date)

    # Move Gottesdienste sheet to the far right
    wb.move_sheet("Gottesdienste", offset=len(wb.sheetnames) - 1)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def send_email(rows, kirchenjahr_rows):
    # Send the email with the report
    # Create the container (outer) email message.
    # msg = MIMEMultipart()
    msg = EmailMessage()
    msg['Subject'] = 'Liste der Gottsdienste'
    msg['From'] = smtp_username
    msg['To'] = email_recipients
    msg.set_content("""*** Dies ist eine automatisch generierte Nachricht. ***\n\n
Hallo,\n\nIm Anhang ist die Liste der Gottesdienste.\n\nViele Grüße,\nFrank""")
    excel_bytes = create_excel(rows, kirchenjahr_rows)
    msg.add_attachment(excel_bytes, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='Gottesdienste.xlsx')
    # Send the email
    try:
        # Connect to the SMTP server
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Upgrade the connection to secure
            server.login(smtp_username, SMTP_PASSWORD)  # Log in to the SMTP server
            server.send_message(msg)  # Send the email
            print("Email sent successfully!")
    except Exception as e:
        print(f"Failed to send email: {e}")

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--from", dest="from_date", default=None, help="Start date YYYY-MM-DD")
    parser.add_argument("--to",   dest="to_date",   default=None, help="End date YYYY-MM-DD")
    args = parser.parse_args()
    rows, kirchenjahr_rows = get_services(from_date=args.from_date, to_date=args.to_date)
    print(f"{len(rows) - 1} Gottesdienste gefunden.")
    if use_email:
        send_email(rows, kirchenjahr_rows)

if __name__ == "__main__":
    main()